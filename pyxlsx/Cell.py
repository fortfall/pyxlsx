import logging
from openpyxl.cell import Cell as _Cell

logger = logging.getLogger(__name__)

class Cell(_Cell):
    __slots__ = ('_cache', '_cache_type')
    def __init__(self, worksheet, row=None, column=None, value=None, style_array=None, cache=None, cache_type=None):
        super().__init__(worksheet, row=row, column=column, value=value, style_array=style_array)
        self._cache = cache
        self._cache_type = cache_type
    
    @property
    def read_only(self):
        return self.parent.read_only

    @property
    def is_formula(self):
        return self.data_type == 'f'
    
    @property
    def cache(self):
        return self._cache
    
    @property
    def cache_type(self):
        return self._cache_type

    @property
    def value(self):
        """Get or set the value held in the cell.

        :type: depends on the value (string, float, int or
            :class:`datetime.datetime`)
        """
        return self._value

    @value.setter
    def value(self, value):
        if self.read_only and self._value is not None:
            raise AttributeError("Cell is read only")
        self._value = value

    @property
    def data(self):
        # out = self._value
        if self.is_formula:
            if self.cache is not None:
                out = self.cache
            else:
                out = self.parent.parent._compute(self.parent.title, self.coordinate)
            if out == None:
                out = self._value
        else:
            out = self._value
        if self.parent.header is None or self.row <= self.parent.header_row:
            return out
        if self.parent.use_default:
            if out is None:
                try:
                    out = self.parent.header.get_default(self.column)
                except Exception as e:
                    logger.debug(str(e))
            else:
                default_type = self.parent.header.get_type(self.column)
                if default_type is not None:
                    try:
                        out = default_type(out)
                    except Exception as e:
                        if default_type == int:
                            out = 0
                        elif default_type == float:
                            out = 0.0
                        logger.warning(str(e))
        return out
    
    @data.setter
    def data(self, value):
        if self.read_only and self._value is not None:
            raise AttributeError("Cell is read only")
        self._bind_value(value)

    @property
    def horizontal(self):
        return tuple(self.parent.cell(self.row, x) for x in range(self.column, self.parent.max_column + 1))
    
    @horizontal.setter
    def horizontal(self, it):
        for idx, item in enumerate(it):
            self.parent.cell(self.row, self.column + idx, item)
    
    @property
    def vertical(self):
        return tuple(self.parent.cell(x, self.column) for x in range(self.row, self.parent.max_row + 1))
    
    @vertical.setter
    def vertical(self, it):
        for idx, item in enumerate(it):
            self.parent.cell(self.row + idx, self.column, item)
    
    @property
    def top(self):
        if self.row == 1:
            raise IndexError(f"cell({self.row}, {self.column}) has no cell above.")
        return self.parent.cell(self.row - 1, self.column)
    
    @property
    def bottom(self):
        return self.parent.cell(self.row + 1, self.column)
    
    @property
    def left(self):
        if self.column == 1:
            raise IndexError(f"cell({self.row}, {self.column}) has not cell on its left.")
        return self.parent.cell(self.row, self.column - 1)
    
    @property
    def right(self):
        return self.parent.cell(self.row, self.column + 1)
    
    def _bind_value(self, value):
        # clear cached formula result
        if self.is_formula and value != self._value:
            self._cache = None
            self._cache_type = None
        super()._bind_value(value)
    

        
            
            


            
