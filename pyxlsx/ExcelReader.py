import warnings
from io import BytesIO
from zipfile import ZipFile, ZIP_DEFLATED

from openpyxl.xml.constants import COMMENTS_NS
from openpyxl.cell import MergedCell
from openpyxl.comments.comment_sheet import CommentSheet
from openpyxl.packaging.relationship import (
    RelationshipList,
    get_dependents,
    get_rels_path
)

from openpyxl.pivot.table import TableDefinition
from openpyxl.chartsheet import Chartsheet
from openpyxl.worksheet.table import Table
from openpyxl.drawing.spreadsheet_drawing import SpreadsheetDrawing
from openpyxl.xml.constants import (
    ARC_SHARED_STRINGS,
    ARC_CORE,
    ARC_CONTENT_TYPES,
    ARC_WORKBOOK,
    ARC_THEME,
    COMMENTS_NS,
    SHARED_STRINGS,
    EXTERNAL_LINK,
    XLTM,
    XLTX,
    XLSM,
    XLSX,
)
from openpyxl.xml.functions import fromstring
from openpyxl.reader.drawings import find_images

from openpyxl.reader.excel import ExcelReader as _ExcelReader, _find_workbook_part
from .WorksheetReader import WorksheetReader
from .WorkbookParser import WorkbookParser
'''
data_only default to False
'''
class ExcelReader(_ExcelReader):
    def read_workbook(self):
        wb_part = _find_workbook_part(self.package)
        self.parser = WorkbookParser(self.archive, wb_part.PartName[1:], keep_links=self.keep_links)
        self.parser.parse()
        wb = self.parser.wb
        wb._sheets = []
        wb._data_only = self.data_only
        wb._read_only = self.read_only
        wb.template = wb_part.ContentType in (XLTX, XLTM)

        # If are going to preserve the vba then attach a copy of the archive to the
        # workbook so that is available for the save.
        if self.keep_vba:
            wb.vba_archive = ZipFile(BytesIO(), 'a', ZIP_DEFLATED)
            for name in self.valid_files:
                wb.vba_archive.writestr(name, self.archive.read(name))

        if self.read_only:
            wb._archive = self.archive

        self.wb = wb

    def read_worksheets(self):
        comment_warning = """Cell '{0}':{1} is part of a merged range but has a comment which will be removed because merged cells cannot contain any data."""
        for sheet, rel in self.parser.find_sheets():
            if rel.target not in self.valid_files:
                continue

            if "chartsheet" in rel.Type:
                self.read_chartsheet(sheet, rel)
                continue

            rels_path = get_rels_path(rel.target)
            rels = RelationshipList()
            if rels_path in self.valid_files:
                rels = get_dependents(self.archive, rels_path)

            # Used Worksheet and WorksheetReader for all scenarios, 
            # not distinguishing between read_only or not
            fh = self.archive.open(rel.target)
            ws = self.wb.create_sheet(sheet.name)
            ws._rels = rels
            ws_parser = WorksheetReader(ws, fh, self.shared_strings, self.data_only, self.read_only)
            ws_parser.bind_all()

            # assign any comments to cells
            for r in rels.find(COMMENTS_NS):
                src = self.archive.read(r.target)
                comment_sheet = CommentSheet.from_tree(fromstring(src))
                for ref, comment in comment_sheet.comments:
                    try:
                        ws[ref].comment = comment
                    except AttributeError:
                        c = ws[ref]
                        if isinstance(c, MergedCell):
                            warnings.warn(comment_warning.format(ws.title, c.coordinate))
                            continue

            # preserve link to VML file if VBA
            if self.wb.vba_archive and ws.legacy_drawing:
                ws.legacy_drawing = rels[ws.legacy_drawing].target

            for t in ws_parser.tables:
                src = self.archive.read(t)
                xml = fromstring(src)
                table = Table.from_tree(xml)
                ws.add_table(table)

            drawings = rels.find(SpreadsheetDrawing._rel_type)
            for rel in drawings:
                charts, images = find_images(self.archive, rel.target)
                for c in charts:
                    ws.add_chart(c, c.anchor)
                for im in images:
                    ws.add_image(im, im.anchor)

            pivot_rel = rels.find(TableDefinition.rel_type)
            for r in pivot_rel:
                pivot_path = r.Target
                src = self.archive.read(pivot_path)
                tree = fromstring(src)
                pivot = TableDefinition.from_tree(tree)
                pivot.cache = self.parser.pivot_caches[pivot.cacheId]
                ws.add_pivot(pivot)

            ws.sheet_state = sheet.state