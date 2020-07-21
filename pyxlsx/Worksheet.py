import logging
from typing import Generator, Union, Optional
from inspect import isgenerator
from openpyxl.worksheet.worksheet import Worksheet as _Worksheet
from openpyxl.utils import (
    coordinate_to_tuple,
    column_index_from_string
)
from openpyxl.worksheet.views import (
    Pane,
    Selection,
    SheetViewList,
)
from .Series import ContentRow, Header, InvalidOperationError
from .Cell import Cell

logging.basicConfig()
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

class Worksheet(_Worksheet):
    def __init__(self, parent, title=None):
        super().__init__(parent, title)
        self._header_row: Optional[int] = None
        self.header: Header = None
    
    @property
    def read_only(self):
        return self.parent._read_only
    
    @property
    def header_row(self):
        '''
        Which row is the header row. When header_row is not None, header is initialized.
        '''
        return self._header_row
    
    @header_row.setter
    def header_row(self, value: int):
        '''
        set the row of header and init worksheet's header
        '''
        if not isinstance(value ,int) and value is not None:
            raise TypeError(f"header_row should be int or None. (got {value}: {type(value)})")
        if isinstance(value, int) and value <= 0:
            raise ValueError(f"header_row can only be >= 1. (got {value})")
        self._header_row = value
        if value is None:
            self.header = None
        else:
            self._init_header(value)
    
    def _init_header(self, header_row):
        self.header = Header(self, self[header_row])

    def insert_rows(self, idx, amount=1):
        '''
        Insert amount of rows before the idx-th row.
        Args:
            idx: the idx-th row
            amount: amount of rows to be inserted
        '''
        if amount == 0:
            return
        super().insert_rows(idx, amount)
        if self.header is not None and idx <= self.header_row:
            self.header_row += amount
            self.header.refresh(new_row=self.header_row)
    
    def delete_rows(self, idx, amount=1):
        '''
        Delete amount of rows from the idx-th row (inclusive).
        Args:
            idx: the idx-th row
            amount: amount of rows to be deleted
        '''
        if amount == 0:
            return
        super().delete_rows(idx, amount)
        if self.header is None or self.header_row < idx:
            return
        elif self.header_row >= idx and self.header_row < idx + amount:
            self.header_row = None
        elif self.header_row >= idx + amount:
            self.header_row -= amount
            self.header.refresh(new_row=self.header_row)
    
    def insert_cols(self, idx, amount=1):
        '''
        Insert amount of columns before the idx-th column.
        Args:
            idx: the idx-th column
            amount: amount of columns to be inserted
        '''
        if type(idx) != int:
            raise TypeError(f"idx should be of type int (got {type(idx)}.")
        if type(amount) != int:
            raise TypeError(f"idx should be of type int (got {type(amount)}.")
        if idx <= 0:
            raise ValueError(f"Cannot insert before the {idx} row.")
        if amount < 0:
            raise ValueError(f"Cannot insert {amount} rows.")
        if amount == 0:
            return
        super().insert_cols(idx, amount)
        self.header.rebuild()

    def delete_cols(self, idx, amount=1):
        '''
        Delete amount of columns from the idx-th column (inclusive).
        Args:
            idx: the idx-th column
            amount: amount of columns to be deleted
        '''
        if type(idx) != int:
            raise TypeError(f"idx should be of type int (got {type(idx)}.")
        if type(amount) != int:
            raise TypeError(f"idx should be of type int (got {type(amount)}.")
        if idx <= 0:
            raise ValueError(f"Cannot insert before the {idx} row.")
        if amount < 0:
            raise ValueError(f"Cannot insert {amount} rows.")
        if amount == 0:
            return
        super().delete_cols(idx, amount)
        self.header.rebuild()
    
    def append_header(self, iterable):
        '''
        Append iterable as header row to worksheet (append a row and set sheet's row_offset).
        '''
        self.append(iterable)
        self.header_row = self._current_row
    
    def append_by_header(self, data_dict, update_header=True):
        '''
        Append to worksheet by header names. Header is auto-generated from dict keys the first time append_by_header is called.
        Args:
            data_dict: a dict containing the row data
            append_header: if True, append new header names from data_dict; else the key-value pair is ignored.
        '''
        if not isinstance(data_dict, dict):
            raise TypeError(f"append_by_header accepts only dict var (got {type(data_dict)}.")
        if self.header is None:
            names = [k for k in data_dict]
            self.append_header(names)
        converted = {}
        for k, v in data_dict.items():
            column = self.header.get_column(k)
            if column is not None:
                converted[column] = v
            elif update_header:
                self.header.append(k)
                converted[self.header.max_column] = v
        self.append(converted)
    
    @property
    def content_rows(self) -> Generator[ContentRow, None, None]:
        '''
        Return a generator of ContentRow.
        '''
        if self.header is None:
            raise InvalidOperationError(f"Cannot init content_rows when header is None. Set header_row to init header.")
        for row in self.iter_rows(self.header_row + 1):
            yield ContentRow(self, row)
    
    def get_content_column(self, key):
        '''
        Get a ContentColumn by column if key is of type int, by header if key is of type str.
        '''
        if self.header is None:
            raise InvalidOperationError("Cannot get content column when header is None. Set header_row to init header.")
        return self.header.cell(key).bottom.vertical

    def _get_cell(self, row, column):
        """
        Internal method for getting a cell from a worksheet.
        Will create a new cell if one doesn't already exist.
        """
        coordinate = (row, column)
        if not coordinate in self._cells:
            cell = Cell(self, row=row, column=column)
            self._add_cell(cell)
            return cell
        return self._cells[coordinate]

    @property
    def freeze_panes(self):
        if self.sheet_view.pane is not None:
            return self.sheet_view.pane.topLeftCell

    @freeze_panes.setter
    def freeze_panes(self, topLeftCell=None):
        if isinstance(topLeftCell, Cell):
            topLeftCell = topLeftCell.coordinate
        if topLeftCell == 'A1':
            topLeftCell = None

        if not topLeftCell:
            self.sheet_view.pane = None
            return

        row, column = coordinate_to_tuple(topLeftCell)

        view = self.sheet_view
        view.pane = Pane(topLeftCell=topLeftCell,
                        activePane="topRight",
                        state="frozen")
        view.selection[0].pane = "topRight"

        if column > 1:
            view.pane.xSplit = column - 1
        if row > 1:
            view.pane.ySplit = row - 1
            view.pane.activePane = 'bottomLeft'
            view.selection[0].pane = "bottomLeft"
            if column > 1:
                view.selection[0].pane = "bottomRight"
                view.pane.activePane = 'bottomRight'

        if row > 1 and column > 1:
            sel = list(view.selection)
            sel.insert(0, Selection(pane="topRight", activeCell=None, sqref=None))
            sel.insert(1, Selection(pane="bottomLeft", activeCell=None, sqref=None))
            view.selection = sel

    def append(self, iterable):
        """Appends a group of values at the bottom of the current sheet.

        * If it's a list: all values are added in order, starting from the first column
        * If it's a dict: values are assigned to the columns indicated by the keys (numbers or letters)

        :param iterable: list, range or generator, or dict containing values to append
        :type iterable: list|tuple|range|generator or dict

        Usage:

        * append(['This is A1', 'This is B1', 'This is C1'])
        * **or** append({'A' : 'This is A1', 'C' : 'This is C1'})
        * **or** append({1 : 'This is A1', 3 : 'This is C1'})

        :raise: TypeError when iterable is neither a list/tuple nor a dict

        """
        row_idx = self._current_row + 1

        if (isinstance(iterable, (list, tuple, range))
            or isgenerator(iterable)):
            for col_idx, content in enumerate(iterable, 1):
                if isinstance(content, Cell):
                    # compatible with write-only mode
                    cell = content
                    if cell.parent and cell.parent != self:
                        raise ValueError("Cells cannot be copied from other worksheets")
                    cell.parent = self
                    cell.column = col_idx
                    cell.row = row_idx
                else:
                    cell = Cell(self, row=row_idx, column=col_idx, value=content)
                self._cells[(row_idx, col_idx)] = cell

        elif isinstance(iterable, dict):
            for col_idx, content in iterable.items():
                if isinstance(col_idx, str):
                    col_idx = column_index_from_string(col_idx)
                cell = Cell(self, row=row_idx, column=col_idx, value=content)
                self._cells[(row_idx, col_idx)] = cell

        else:
            self._invalid_row(iterable)

        self._current_row = row_idx

